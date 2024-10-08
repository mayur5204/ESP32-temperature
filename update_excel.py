import json
from openpyxl import load_workbook
from datetime import datetime

# Load data from the JSON file fetched from Adafruit IO
with open('adafruit_data.json') as f:
    data = json.load(f)[0]  # Get the first (latest) entry

# Load or create the Excel workbook
file_name = 'data.xlsx'

try:
    workbook = load_workbook(file_name)
    sheet = workbook.active
except FileNotFoundError:
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    # Create headers if the file doesn't exist
    sheet.append(['Timestamp', 'Feed Key', 'Value', 'Latitude', 'Longitude', 'Elevation'])

# Prepare data to append
timestamp = data['created_at']
feed_key = data['feed_key']
value = data['value']
lat = data.get('lat', 'N/A')  # Default to 'N/A' if lat is missing
lon = data.get('lon', 'N/A')  # Default to 'N/A' if lon is missing
ele = data.get('ele', 'N/A')  # Default to 'N/A' if ele is missing

# Append the new row
sheet.append([timestamp, feed_key, value, lat, lon, ele])

# Save the updated Excel file
workbook.save(file_name)
